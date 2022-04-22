from openpyxl import load_workbook
import os
import re

join = os.path.join
working_dir = os.path.normpath("C:/Users/SQA-AGrudev/Desktop/TAF_WORKING_DIR")
working_dir_json = join(working_dir, 'Jsons')
working_dir_xlsx = join(working_dir, 'XLSX')
working_dir_sorted = join(working_dir, 'Sorted')

os.chdir(working_dir)


def rename_files(name_to_be_ranamed):
    """
    Renames files in the current working directory directory to match the current convention.
    """
    os.chdir(working_dir)
    working_name = str(name_to_be_ranamed)
    if re.search("PostPaid_CON", working_name):
        working_name = re.sub('PostPaid_CON', '', working_name)
        working_name = re.sub('_', "_PostPaid_CON_", working_name, 1)
    if re.search("PostPaid_BUS", working_name):
        working_name = re.sub('PostPaid_BUS', '', working_name)
        working_name = re.sub('_', "_PostPaid_BUS_", working_name, 1)
    if re.search("PostPaid_SP", working_name):
        working_name = re.sub('PostPaid_SP', '', working_name)
        working_name = re.sub('_', "_PostPaid_SP_", working_name, 1)
    if re.search("_LTE", working_name):
        working_name = re.sub('_LTE', '', working_name)
        working_name = re.sub('_', "_LTE_", working_name, 1)
    working_name = re.sub('Terminate', 'Delete', working_name)
    working_name = re.sub('Register', 'Create', working_name)
    working_name = re.sub('04947', '', working_name)
    working_name = re.sub('04451', '', working_name)
    working_name = re.sub('04452', '', working_name)
    working_name = re.sub('04166', '', working_name)
    working_name = re.sub('04946', '', working_name)
    working_name = re.sub('04334', '', working_name)
    working_name = re.sub('04460', '', working_name)
    working_name = re.sub('__', '_', working_name)
    working_name = re.sub('_\s', '_', working_name)
    working_name = re.sub('_\.', '.', working_name)
    working_name = re.sub('Change_IMSI', 'ChangeIMSI', working_name)
    print(working_name)
    os.rename(name_to_be_ranamed, working_name)


def operate_cwd_old(work_on):
    """
    Operates over the files in the given working directory comment and uncomment needed functions
    returns a list of file names for later sorting
    Also returns a list of files later use for numbering them
    """
    # JSON
    if work_on == 'json':
        for root, dirs, files in os.walk(working_dir_json, topdown=False):
            if root != working_dir_json:
                continue
            for name in files:
                if name[-4:] == 'json':
                    pass
                    remove_request_trace(name)
                    get_service_names(name)
                    replace_service_names(name)
                    if os.path.exists(join(working_dir_json, "service_names.txt")):
                        os.replace(join(working_dir_json, "service_names.txt"), join(working_dir, "service_names.txt"))
                    parametrize_payload(name)
    if work_on == 'json':
        for root, dirs, files in os.walk(working_dir_json, topdown=False):
            if root != working_dir_json:
                continue
            for name in files:
                if name[-4:] == 'json':
                    replace_var_value(False, name)
    # XLS
    if work_on == 'xlsx':
        for root, dirs, files in os.walk(working_dir_xlsx, topdown=False):
            if root != working_dir_xlsx:
                continue
            for name in files:
                if name[-4:] == 'xlsx':
                    pass
                    workbook = load_workbook(filename=join(root, name))
                    ws = load_sheet(workbook)
                    if not ws:
                        print(f'Failed to load sheet for file {name}')
                        break
                    remove_resp_add_first_four_lines(ws)
                    update_names(ws, name[:-4] + 'json')
                    remove_sri(ws)
                    remove_antics(ws)
                    replace_var_value(ws, name)
                    workbook.save(filename=join(root, name))
                print(f"Successfully modified File:{name}")


def insert_in_nvar(size_to_match, string_to_insert, modified_list):
    """
    Fills the modified list with as many values as the size of the first list
    """
    for _ in size_to_match:
        modified_list = modified_list + [string_to_insert]
    return modified_list

def read_params_from_file(file_name,list_tmp):
    with open(file_name, 'r') as file:
        for line in file:
            line = line.strip('\n')
            line = f'{line}'
            list_tmp.append(line)
    return list_tmp
def replace_var_value(ws, name_working):
    """
    Parametrazis the excels and json files.
    """
    ICCID = []
    IMSI = []
    order_subsceriber_desc = []
    order_subscriber_id = []
    service_name = []
    os.chdir(join(working_dir, 'configuration'))
    order_subsceriber_desc=read_params_from_file('order_subscriber_desc.txt',order_subsceriber_desc)
    order_subscriber_id=read_params_from_file('order_subscriber_id.txt', order_subscriber_id)
    ICCID = read_params_from_file('iccid.txt', ICCID)
    IMSI = read_params_from_file('imsi.txt', IMSI)
    service_name = read_params_from_file('service_names.txt', service_name)
    system_component_id = ['pcrf_1384193797', "pcrf_1084193793", "RoamingZone", ]
    valid_to_date = ['05-09-2022']
    roaming_zone = ['RZ1']

    o_var = order_subsceriber_desc + order_subscriber_id + service_name + system_component_id + valid_to_date + roaming_zone + ICCID + IMSI

    n_var = []

    n_var = insert_in_nvar(order_subsceriber_desc, '${ORDER_SUBSCRIBER_DESC}', n_var)
    n_var = insert_in_nvar(order_subscriber_id, '${ORDER_SUBSCRIBER_ID}', n_var)
    n_var = insert_in_nvar(service_name, '${SERVICE_NAME}', n_var)
    n_var = insert_in_nvar(system_component_id, '${SYSTEM_COMPONENT_ID}', n_var)
    n_var = insert_in_nvar(valid_to_date, "${VALID_TO_DATE}", n_var)
    n_var = insert_in_nvar(roaming_zone, "${ROAMING_ZONE}", n_var)
    n_var = insert_in_nvar(ICCID, "${ICCID}", n_var)
    n_var = insert_in_nvar(IMSI, "${IMSI}", n_var)
    if name_working[-4:] == 'json':
        replace_in_json(name_working, o_var, n_var)
    elif name_working[-4:] == 'xlsx':
        replace_in_xls(ws, o_var, n_var)


def replace_in_json(name_working, o_var, n_var):
    """
    Parametrize the JSON file
    """
    with open(join(working_dir_json, name_working), "r") as file:
        f = file.read()
        for idx, _ in enumerate(o_var):
            f = re.sub(o_var[idx], n_var[idx], f)
    with open(join(working_dir_json, name_working), "w") as file:
        file.write(f)


def replace_in_xls(ws, old_var, new_var):
    """
    Parametrize the XLS file
    """
    if not ws:
        return
    collum = 'J'
    for idx, _ in enumerate(old_var):
        i = 8
        while ws[collum + str(i)].value is not None:
            ws[collum + str(i)].value = re.sub(old_var[idx], new_var[idx], ws[collum + str(i)].value)
            if ws[collum + str(i)].value == old_var[idx]:
                ws[collum + str(i)] = new_var[idx]
            i += 1


def remove_sri(ws):
    """
    Removes the SRI tasks.
    """
    collum_key = 'I'
    collum_value = 'J'
    collum_task = 'K'
    i = 2
    f = 0
    while ws[collum_value + str(i)].value is not None:
        if (ws[collum_value + str(i)].value == 'SRI') and (ws[collum_key + str(i)].value == 'NE_TYPE'):
            task_id = ws[collum_task + str(i)].value
            while ws[collum_task + str(i)].value == task_id:
                ws.delete_rows(i, 1)
                f = 1
            if f == 1:
                i -= 1
        i += 1


def remove_resp_add_first_four_lines(ws):
    """
    Removes the response parameters and add additional 4 lines to each excel file.
    """
    if ws["I4"].value != "TimeOut":
        ws.insert_rows(4)
        ws.insert_rows(4)
        ws.insert_rows(4)
        ws.insert_rows(4)

        ws['D' + str(4)].value = 'IL'
        ws['E' + str(4)].value = 'FlowoneAPI'
        ws['F' + str(4)].value = 'Setup'
        ws['G' + str(4)].value = 'AsyncReceiver'
        ws['I' + str(4)].value = 'TimeOut'
        ws['J' + str(4)].value = '${TimeOut}'

        ws['D' + str(5)].value = 'IL'
        ws['E' + str(5)].value = 'FlowoneAPI'
        ws['F' + str(5)].value = 'Verify'
        ws['G' + str(5)].value = 'AckResponse'
        ws['I' + str(5)].value = 'statusMessage'
        ws['J' + str(5)].value = 'InstantLink accepted request with request id: ${Request_id} for order no: 1234567890'

        ws['D' + str(6)].value = 'IL'
        ws['E' + str(6)].value = 'FlowoneAPI'
        ws['F' + str(6)].value = 'Verify'
        ws['G' + str(6)].value = 'Response'
        ws['H' + str(6)].value = 'serviceOrder'
        ws['I' + str(6)].value = 'statusMessage'
        ws['J' + str(6)].value = 'Order delivered'
        ws['K' + str(6)].value = 'Resp'

        ws['D' + str(7)].value = 'IL'
        ws['E' + str(7)].value = 'FlowoneAPI'
        ws['F' + str(7)].value = 'Verify'
        ws['G' + str(7)].value = 'Response'
        ws['H' + str(7)].value = 'serviceOrder'
        ws['I' + str(7)].value = 'state'
        ws['J' + str(7)].value = 'completed'
        ws['K' + str(7)].value = 'Resp'

    collum_value = 'H'
    i = 8
    while ws['G' + str(i)].value is not None:
        if ws[collum_value + str(i)].value == 'TaskResponse':
            ws.delete_rows(i, 1)
        if ws['G' + str(i)].value == 'total_ne_tasks':
            ws.delete_rows(i, 1)
        i += 1


def update_names(ws, name_to_be_updated):
    """
    Upadate the test name in the excel to match the name of the excel file.
    Also gives the same name to the json that is required as input.
    """
    collum_value = 'I'
    i = 2
    while ws[collum_value + str(i)].value is not None:
        if ws["H" + str(i)].value == "RequestPayload":
            ws["I" + str(i)].value = "PayloadFile"
            ws["J" + str(i)].value = name_to_be_updated
        i += 1
    ws['A3'].value = name_to_be_updated[:-5]
    ws['B3'].value = name_to_be_updated[:-5]


def remove_antics(ws):
    """
    Removes duplicate variables in the excel files
    """
    collum_key = 'I'
    collum_value = 'J'
    collum_task = 'K'
    i = 2
    while ws[collum_value + str(i + 1)].value is not None:
        task_id = ws[collum_task + str(i)].value
        param_value = ws[collum_key + str(i)].value
        j = i + 1
        while task_id == ws[collum_task + str(j)].value:
            if param_value == ws[collum_key + str(j)].value:
                ws.delete_rows(j, 1)
            j += 1
        i += 1


def sort_files(work_on):
    """
    Sorts the payload files into Subscribers
    """
    wd = False
    groups = ["PostPaid_CON", "PostPaid_BUS", "PostPaid_SP", "LTE", "PrePaid_SP", "PrePaid", "M2M", "IPT", "VOICETWIN"]

    if work_on == 'json':
        wd = working_dir_json
    if work_on == 'xlsx':
        wd = working_dir_xlsx

    for root, dirs, files in os.walk(wd, topdown=False):
        if root != wd:
            continue
        for name in files:
            os.chdir(wd)
            for group in groups:
                if re.search(group, name):
                    try:
                        os.replace(join(wd, name), join(working_dir_sorted, group, name))
                    except:
                        pass


def give_number_to_files(starting_number):
    """
    Renames files in the current working directory directory to be sorted in order and in create delete pairs
    """

    for root, dirs, files in os.walk(working_dir, topdown=False):
        file_names = files
        k = 0
        os.chdir(root)
        i = starting_number
        if root == working_dir:
            while k < len(file_names) and i < len(file_names) + starting_number:
                j = 0
                while j < len(file_names):
                    if file_names[k][6:] == file_names[j][6:] and k != j:
                        try:
                            if i < 9:
                                os.rename(file_names[k], '000' + str(i) + file_names[k])
                                os.rename(file_names[j], '000' + str(i + 1) + file_names[j])
                            elif i == 9:
                                os.rename(file_names[k], '000' + str(i) + file_names[k])
                                os.rename(file_names[j], '00' + str(i + 1) + file_names[j])
                            else:
                                os.rename(file_names[k], '00' + str(i) + file_names[k])
                                os.rename(file_names[j], '00' + str(i + 1) + file_names[j])
                            i += 2
                        except OSError as err:
                            print(err)
                        break
                    j += 1
                k += 1


def remove_number_from_files():
    """
    Renames files in the current working directory directory to remove the sorting numbers
    """
    os.chdir(working_dir)
    file_names = []
    for root, dirs, files in os.walk(working_dir, topdown=False):
        file_names = files
        if root == working_dir:
            break
    k = 0
    while k < len(file_names):
        if re.match('\d{4}', file_names[k]) is not None:
            os.rename(file_names[k], file_names[k][4:])
        k += 1


def remove_request_trace(name_working):
    """
    Removes the request trace part  from the json files.
    """
    with open(join(working_dir_json, name_working), "r") as file:
        f = file.read()
        f = re.sub(',\s*{\s*"name":\s"REQUEST_TRACE",\s*"value":\s"EVERYTHING"\s*}', '', f)
    with open(join(working_dir_json, name_working), "w") as file:
        file.write(f)


def replace_service_names(name_working):
    """
    Replaces the Service name value with ${SERVICE_NAME} in all json and does the same for other common parameters
    """

    with open(join(working_dir_json, name_working), "r") as file:
        file_string = file.read()
        service_match = re.search(r'"service":\s{\s*"id":\s"([a-zA-z0-9]*)",\s*"name":\s"([0-9]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"service":\s{\s*"id":\s"([a-zA-z0-9]*)",\s*"name":\s"([0-9]*)"',
                                 '\t"service": {\n        "id":"%s",\n        \t"name":"${SERVICE_NAME}"' % (
                                     service_match.group(1)), file_string)

        service_match = re.search(
            r'"serviceCharacteristic":\s\[\s*{\s*"name":\s?"116",\s*"value":\s?"([0-9]*)"\s*}\s*]', file_string)
        if service_match is not None:
            file_string = re.sub('"serviceCharacteristic":\s\[\s*{\s*"name":\s?"116",\s*"value":\s?"([0-9]*)"',
                                 '\t\t"serviceCharacteristic": [{\n\t\t\t\t"name":"116",\n\t\t\t\t"value":"${ORDER_SUBSCRIBER_DESC_NEW}"',
                                 file_string)

        service_match = re.search(r'"name":\s?"6407",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"6407",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"6407",\n\t\t\t\t"value":"${MBN_AGREEMENT_ID}"',file_string)

        service_match = re.search(r'"name":\s?"6875",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"6875",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"6875",\n\t\t\t\t"value":"${KURT_ID}"',file_string)

        service_match = re.search(r'"name":\s?"3086",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"3086",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"3086",\n\t\t\t\t"value":"${MBN_FIRST_NAME}"',file_string)

        service_match = re.search(r'"name":\s?"3087",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"3087",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"3087",\n\t\t\t\t"value":"${MBN_SECOND_NAME}"', file_string)

        service_match = re.search(r'"name":\s?"6304",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"6304",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"6304",\n\t\t\t\t"value":"${MBN_EMAIL}"',file_string)

        service_match = re.search(r'"name":\s?"6493",\s*"value":\s?"([^"]*)"', file_string)
        if service_match is not None:
            file_string = re.sub('"name":\s?"6493",\s*"value":\s?"([^"]*)"',
                                 '\t\t"name":"6493",\n\t\t\t\t"value":"${MBN_FLAG}"', file_string)
    with open(join(working_dir_json, name_working), "w") as file:
        file.write(file_string)


def get_service_names(file_name):
    """
    Reads the service name from the json files and prints them to the screen.
    """
    os.chdir(working_dir)
    with open(join(working_dir_json, file_name), "r") as file:
        file_string = file.read()
        service_match = re.search(r'"service":\s{\s*"id":\s"([a-zA-z0-9]*)",\s*"name":\s"([0-9]*)"', file_string)

    if service_match is not None:
        print(f',"{service_match.group(2)}"')
        with open('service_names.txt', 'a') as file:
            file.write(f'{service_match.group(2)}\n')


def parametrize_payload(name_working):
    """
    Parametrizes the payload with the default values for the subscriber it dose not replace the values for the service
    """
    with open(join(working_dir_json, name_working), "r") as file:
        file_string = file.read()
        file_string = re.sub('"externalId":\s?"([^"]*)"', '"externalId": "${externalId}"', file_string)
        file_string = re.sub('"orderDate":\s?"([^"]*)"', '"orderDate": "${orderDate}"', file_string)
        file_string = re.sub('"orderType":\s?"([^"]*)"', '"orderType": "${orderType}"', file_string)

        match = re.search(r'"AsyncResponse":\s??{[\s\S]*"replyToAddress":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${AsyncResponse_replyToAddress}', file_string)

        match = re.search(r'"WSSec":\s?{[\s\S]*"username":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${WSSec_username}', file_string)

        match = re.search(r'"WSSec":\s?{[\s\S]*"password":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${WSSec_password}', file_string)

        match = re.search(r'"OMRequestSpec":\s?{[\s\S]*"neType":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${OMRequestSpec_neType}', file_string)

        match = re.search(r'"name":\s?"IL_REQ_GROUP",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${IL_REQ_GROUP}', file_string)

        match = re.search(r'"name":\s?"ORDER_SUBSCRIBER_ID",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${ORDER_SUBSCRIBER_ID}', file_string)

        match = re.search(r'"name":\s?"ORDER_SUBSCRIBER_DESC",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${ORDER_SUBSCRIBER_DESC}', file_string)

        match = re.search(r'"name":\s?"ORDER_SYS_ID",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${ORDER_SYS_ID}', file_string)

        match = re.search(r'"name":\s?"CASE_ID",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${CASE_ID}', file_string)

        match = re.search(r'"name":\s?"ORDER_USER_ID",\s*"value":\s?"([^"]*)"', file_string)
        file_string = re.sub(match.group(1), '${ORDER_USER_ID}', file_string)

    with open(join(working_dir_json, name_working), "w") as file:
        file.write(file_string)


def get_payload(sheet_name, workbook):
    """
    Reads the Dayly report xcel file and creates payload from it. It requres a name collum to be set up at collum C.
    """
    ws = workbook[sheet_name]
    i = 3
    os.chdir(working_dir_json)
    while ws['F' + str(i)].value is not None:
        match = re.search(
            r'[jJ][sS][oO][nN][\s.]?[Rr][eE][Qq][uUeEsStT.]{0,4}:?\s{0,3}([\S\s]*)[Jj][Ss][Oo][Nn][\s.]{0,3}[Rr][Ee][Ss][pPoOnNsSeE.]{0,5}:?',
            ws['F' + str(i)].value)
        test_name = ws['C' + str(i)].value
        if test_name is None:
            print(f'Line {i} Error:{ws["G" + str(i)].value}')
        elif match is None:
            print(f'Payload missing for line {i}')
        else:
            with open(test_name + '.json', 'w') as file:
                file.write(match.group(1))
        i += 1


def get_request_id_cfg_file(sheet_name, workbook):
    """
    Reads the Dayly report xcel file creates the reqest_ids file for the task generator
    and the cfg file for the excel generator
    """
    ws = workbook[sheet_name]
    i = 3
    os.chdir(working_dir)
    request_ids = []
    test_cfg = []
    while ws['F' + str(i)].value is not None:
        match = re.match('[Rr][Ee][Qq][UuEeSsTt]*.?\s?[Ii][Dd]\s{0,2}[.:]{0,2}\s{0,4}(\d+)', ws['F' + str(i)].value)
        test_name = ws['C' + str(i)].value
        if test_name is None:
            print(f'Line {i} Error:{ws["G" + str(i)].value}')
        elif match is None:
            print(f'Request missing for line {i}')
        else:
            request_ids.append(match.group(1))
            # test_cfg.append(f'{test_name}')
            test_cfg.append(f'{match.group(1)}=TC_NAME:{test_name},TC_ID:{test_name}')
        i += 1
    with open('request_ids.txt', "a") as file:
        temp_string = ''
        for el in request_ids:
            temp_string += f'{el}\n'
        file.write(temp_string)
    with open('cfg_file.txt', "a") as file:
        temp_string = ''
        for el in test_cfg:
            temp_string += f'{el}\n'
        file.write(temp_string)


def load_sheet(workbook):
    """
    Some excels are generated with different sheetnames
    """
    try:
        ws = workbook['Test cases1']
    except:
        try:
            ws = workbook['Test cases']
        except:
            return
    return ws


def operate_workbook(curret_day, last_day):
    """
    Operates the Test Report file, generates cfg file for TAF regression generator,
    service names files later used for parametraization and request ids file used to get request from DB
    """
    workbook = load_workbook(join(working_dir, 'Test_Report.xlsx'))
    with open('request_ids.txt', "w") as file:
        pass
    with open('cfg_file.txt', "w") as file:
        pass
    while curret_day < last_day:
        get_payload(f'Day {curret_day}', workbook)
        print(f'Payload for Day {curret_day} Finished \n --------------------------------------')
        get_request_id_cfg_file(f'Day {curret_day}', workbook)
        print(f'Request IDS for Day {curret_day} Finished \n --------------------------------------')
        check_payload(f'Day {curret_day}', workbook)
        curret_day += 1


def check_payload(sheet_name, workbook):
    """
    checks if the payload have mismatch between test case name and request and if a sheet has duplicate REQ IDs
    """
    ws = workbook[sheet_name]
    i = 3
    request_ids_temp = []
    while ws['F' + str(i)].value is not None:
        if ws['C' + str(i)].value is not None:
            create_match_payload = re.search(r'Create', ws['F' + str(i)].value)
            delete_match_payload = re.search(r'Delete', ws['F' + str(i)].value)
            create_match_name = re.search(r'Create', ws['C' + str(i)].value)
            delete_match_name = re.search(r'Delete', ws['C' + str(i)].value)
            if create_match_name is not None and create_match_payload is None:
                print(f"error in sheet{sheet_name} line:{i}")
            if delete_match_name is not None and delete_match_payload is None:
                print(f"error in sheet{sheet_name} line:{i}")
        match = re.match('[Rr][Ee][Qq][UuEeSsTt]*.?\s?[Ii][Dd]\s{0,2}[.:]{0,2}\s{0,4}(\d+)', ws['F' + str(i)].value)
        if match is not None:
            request_ids_temp.append(match.group(1))
        i += 1
    seen = set()
    dupes = []
    for x in request_ids_temp:
        if x in seen:
            dupes.append(x)
        else:
            seen.add(x)
    if dupes:
        print(f'{sheet_name} contains these duplicate request ids:{dupes}')


def init_dir():
    """
    Creates the necessary files and dirs to operate
    """
    if not os.path.isdir(working_dir_json):
        os.mkdir(working_dir_json)
    if not os.path.isdir(working_dir_sorted):
        os.mkdir(working_dir_sorted)
    if not os.path.isdir(working_dir_xlsx):
        os.mkdir(working_dir_xlsx)
    l = ['LTE', 'M2M', "PostPaid_BUS", "PostPaid_CON", "PostPaid_SP", "PrePaid", "PrePaid_SP", "IPT"]
    for i in l:
        if not os.path.isdir(join(working_dir_sorted, i)):
            os.mkdir(join(working_dir_sorted, i))
    with open(join(working_dir, 'service_names.txt'), 'w') as f:
        pass
    with open(join(working_dir, 'cfg_file.txt'), 'w') as f:
        pass
    with open(join(working_dir, 'request_ids.txt'), 'w') as f:
        pass


def clear_dir():
    """
    Removes the jsons and xlsx's
    """
    l = [working_dir_xlsx, working_dir_json, working_dir_sorted]
    for top in l:
        for root, dirs, files in os.walk(top, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))



# TEST REPORT OPERATIONS
# init_dir()

# operate_workbook(13,28)

# operate_cwd_old('json')
operate_cwd_old('xlsx')
# sort_files('xlsx')
# give_number_to_files(54)
# sort_files('json')

# clear_dir()
# remove_number_from_files()
